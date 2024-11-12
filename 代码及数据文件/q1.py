import scipy.stats as stats
import openpyxl

# 二项分布模型
def test_B():
    # 创建一个新的工作簿
    workbook = openpyxl.Workbook()

    # 创建第一个工作表
    sheet1 = workbook.active
    sheet1.title = "情形1,假设实际次品率p=12%"
    sheet1.append(['CDF，即P(X<=k)', '样本量n', '临界次品量k'])

    p = 0.12
    for n in range(1,150):    # 枚举样本量
        for k in range(n):  # 枚举临界值
            # 计算二项分布的累积分布函数 P(X <= k)
            cdf_value = stats.binom.cdf(k, n, p)
            if(cdf_value<=0.05):# and cdf_value>=0.04):    # 希望更接近0.05
                #print(cdf_value,n,k)
                sheet1.append([cdf_value,n,k])

    # 创建第一个工作表
    sheet1 = workbook.create_sheet(title="情形1,假设实际次品率p=15%")
    sheet1.append(['CDF，即P(X<=k)', '样本量n', '临界次品量k'])

    p = 0.15
    for n in range(1,150):    # 枚举样本量
        for k in range(n):  # 枚举临界值
            # 计算二项分布的累积分布函数 P(X <= k)
            cdf_value = stats.binom.cdf(k, n, p)
            if(cdf_value<=0.05):# and cdf_value>=0.04):    # 希望更接近0.05
                #print(cdf_value,n,k)
                sheet1.append([cdf_value,n,k])

    # 创建第二个工作表
    sheet2 = workbook.create_sheet(title="情形2,假设实际次品率p=5%")
    sheet2.append(['1-CDF，即P(X>k)', '样本量n', '临界次品量k'])

    p = 0.05
    for n in range(1, 150):  # 枚举样本量
        for k in range(n):  # 枚举临界值
            # 计算二项分布的累积分布函数 P(X <= k)
            cdf_value = stats.binom.cdf(k, n, p)
            if (1 - cdf_value <= 0.1):#and 1 - cdf_value>=0.09):    # 希望更接近0.10
                #print(1-cdf_value, n, k)
                sheet2.append([1-cdf_value, n, k])

    # 创建第二个工作表
    sheet2 = workbook.create_sheet(title="情形2,假设p=8%")
    sheet2.append(['1-CDF，即P(X>k)', '样本量n', '临界次品量k'])

    p = 0.08
    for n in range(1, 150):  # 枚举样本量
        for k in range(n):  # 枚举临界值
            # 计算二项分布的累积分布函数 P(X <= k)
            cdf_value = stats.binom.cdf(k, n, p)
            if (1 - cdf_value <= 0.1):# and 1 - cdf_value>=0.09):    # 希望更接近0.10
                #print(1-cdf_value, n, k)
                sheet2.append([1-cdf_value, n, k])

    # 创建第二个工作表
    sheet2 = workbook.create_sheet(title="情形2,假设p=10%")
    sheet2.append(['1-CDF，即P(X>k)', '样本量n', '临界次品量k'])

    p = 0.10
    for n in range(1, 150):  # 枚举样本量
        for k in range(n):  # 枚举临界值
            # 计算二项分布的累积分布函数 P(X <= k)
            cdf_value = stats.binom.cdf(k, n, p)
            if (1 - cdf_value <= 0.1):# and 1 - cdf_value>=0.09):    # 希望更接近0.10
                #print(1-cdf_value, n, k)
                sheet2.append([1-cdf_value, n, k])

    # 保存文件
    workbook.save('问题1——二项分布.xlsx')

def test_H():
    # 创建一个新的工作簿
    workbook = openpyxl.Workbook()

    # 创建第一个工作表
    sheet1 = workbook.active
    sheet1.title = "情形1,假设实际次品率p=12%"
    sheet1.append(['CDF，即P(X<=k)', '总体数M','总体中次品数D', '样本量n', '临界次品量k'])

    p = 0.12  # 次品率
    for M in range(1, 2001):  # 总体中元素的总数（比如零件总数）
        for n in range(1,M):  # 抽取的样本数
            D = int(M*p)
            flag = False
            for k in range(1,min(n,D)):  # 计算 P(X <= k)
                # 计算超几何分布的累积分布函数 P(X <= k)
                cdf_value = stats.hypergeom.cdf(k, M, D, n)
                if(cdf_value<=0.05):# and cdf_value>=0.04):
                    #print(cdf_value,M,D,n,k)
                    sheet1.append([cdf_value,M,D,n,k])
                    flag = True
                    break
            if flag:
                break

    sheet1 = workbook.create_sheet(title="情形1,假设实际次品率p=15%")
    sheet1.append(['CDF，即P(X<=k)', '总体数M','总体中次品数D', '样本量n', '临界次品量k'])

    p = 0.15  # 次品率
    for M in range(1, 2001):  # 总体中元素的总数（比如零件总数）
        for n in range(1,M):  # 抽取的样本数
            D = int(M*p)
            flag = False
            for k in range(1,min(n,D)):  # 计算 P(X <= k)
                # 计算超几何分布的累积分布函数 P(X <= k)
                cdf_value = stats.hypergeom.cdf(k, M, D, n)
                if(cdf_value<=0.05):# and cdf_value>=0.04):
                    #print(cdf_value,M,D,n,k)
                    sheet1.append([cdf_value,M,D,n,k])
                    flag = True
                    break
            if flag:
                break

    sheet1 = workbook.create_sheet(title="情形2,假设实际次品率p=5%")
    sheet1.append(['CDF，即P(X<=k)', '总体数M', '总体中次品数D', '样本量n', '临界次品量k'])

    p = 0.05  # 次品率
    for M in range(1, 2001):  # 总体中元素的总数（比如零件总数）
        for n in range(1,M):  # 抽取的样本数
            D = int(M*p)
            flag = False
            for k in range(1,min(n,D)):  # 计算 P(X <= k)
                # 计算超几何分布的累积分布函数 P(X <= k)
                cdf_value = stats.hypergeom.cdf(k, M, D, n)
                if(1-cdf_value <= 0.10):# and cdf_value>=0.04):
                    #print(cdf_value,M,D,n,k)
                    sheet1.append([1-cdf_value,M,D,n,k])
                    flag = True
                    break
            if flag:
                break

    sheet1 = workbook.create_sheet(title="情形2,假设实际次品率p=8%")
    sheet1.append(['CDF，即P(X<=k)', '总体数M', '总体中次品数D', '样本量n', '临界次品量k'])

    p = 0.08  # 次品率
    for M in range(1, 2001):  # 总体中元素的总数（比如零件总数）
        for n in range(1,M):  # 抽取的样本数
            D = int(M*p)
            flag = False
            for k in range(1,min(n,D)):  # 计算 P(X <= k)
                # 计算超几何分布的累积分布函数 P(X <= k)
                cdf_value = stats.hypergeom.cdf(k, M, D, n)
                if(1-cdf_value <= 0.10):# and cdf_value>=0.04):
                    #print(cdf_value,M,D,n,k)
                    sheet1.append([1-cdf_value,M,D,n,k])
                    flag = True
                    break
            if flag:
                break

    sheet1 = workbook.create_sheet(title="情形2,假设实际次品率p=10%")
    sheet1.append(['CDF，即P(X<=k)', '总体数M', '总体中次品数D', '样本量n', '临界次品量k'])

    p = 0.1  # 次品率
    for M in range(1, 2001):  # 总体中元素的总数（比如零件总数）
        for n in range(1,M):  # 抽取的样本数
            D = int(M*p)
            flag = False
            for k in range(1,min(n,D)):  # 计算 P(X <= k)
                # 计算超几何分布的累积分布函数 P(X <= k)
                cdf_value = stats.hypergeom.cdf(k, M, D, n)
                if(1-cdf_value <= 0.10):# and cdf_value>=0.04):
                    #print(cdf_value,M,D,n,k)
                    sheet1.append([1-cdf_value,M,D,n,k])
                    flag = True
                    break
            if flag:
                break

    # 保存文件
    workbook.save('问题1——超几何分布.xlsx')

test_B()
test_H()

